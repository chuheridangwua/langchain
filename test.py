from openpyxl import load_workbook

try:
    # 加载Excel文件，使用read_only模式提高大文件读取性能
    wb = load_workbook(filename='大10000excel.xlsx', read_only=True)
    
    # 获取所有工作表名称
    sheet_names = wb.sheetnames
    
    # 检查是否存在工作表
    if not sheet_names:
        result = {"警告": "Excel文件中没有工作表"}
    else:
        # 获取第一个工作表
        sheet = wb[sheet_names[0]]
        
        # 读取前5行数据，跳过空行
        data = []
        row_count = 0
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            # 过滤掉全为空值的行
            if any(cell is not None for cell in row):
                data.append(row)
                row_count += 1
        
        # 准备结果字典
        result = {
            "工作表数量": len(sheet_names),
            "工作表名称": sheet_names,
            "第一个工作表的前5行有效数据": data,
            "实际读取行数": row_count
        }
    
    # 显式关闭工作簿释放资源
    wb.close()

except FileNotFoundError:
    result = {"错误": "指定的Excel文件不存在"}
except Exception as e:
    result = {"错误": f"读取Excel文件时发生异常: {str(e)}"}

print(result)